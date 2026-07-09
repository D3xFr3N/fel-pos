from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field

from sqlalchemy.orm import Session

from app.models import Department, InventoryMovement, Product, Supplier

DEFAULT_SUPPLIER_NAME = "Importado inventario"

COLUMN_ALIASES: dict[str, list[str]] = {
    "sku": ["codigo", "code", "sku", "clave", "id producto"],
    "name": ["descripcion", "nombre", "producto", "articulo"],
    "cost": ["costo", "precio de costo", "precio costo"],
    "price": ["precio de venta", "precio venta", "precio", "p venta", "p. venta"],
    "barcode": ["codigo de barras", "barcode", "cb", "ean", "upc"],
    "department": ["departamento", "depto", "categoria"],
    "supplier": ["proveedor", "vendor"],
    "stock": [
        "cantidad en inventario",
        "inventario",
        "existencia",
        "stock",
        "cantidad",
        "qty",
    ],
    "min_stock": ["minimo", "minimo inventario", "stock minimo", "min stock"],
    "wholesale_price": ["precio de mayoreo", "precio mayoreo", "mayoreo"],
}


@dataclass
class EleventaImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    departments_created: int = 0
    suppliers_created: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_header(value: str) -> str:
    text = (value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_barcode(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip().upper()
    return cleaned or None


def _parse_number(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return default


def _read_text_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    decoded = None
    for encoding in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("No se pudo leer el archivo. Guardalo como CSV UTF-8 o Excel .xlsx.")

    sample = decoded[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("El archivo no tiene encabezados de columnas.")
    headers = [header or "" for header in reader.fieldnames]
    rows = [{key or "": (value or "").strip() for key, value in row.items()} for row in reader]
    return headers, rows


def _read_xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Para importar Excel .xlsx instala openpyxl o guarda el archivo como CSV."
        ) from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        raise ValueError("El archivo Excel esta vacio.")

    headers = [str(cell or "").strip() for cell in header_row]
    rows: list[dict[str, str]] = []
    for row_values in iterator:
        if not row_values or all(value in (None, "") for value in row_values):
            continue
        row_dict: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row_values[index] if index < len(row_values) else ""
            row_dict[header] = "" if value is None else str(value).strip()
        rows.append(row_dict)
    return headers, rows


def parse_eleventa_file(content: bytes, filename: str) -> tuple[list[str], list[dict[str, str]]]:
    lowered = (filename or "").lower()
    if lowered.endswith(".xlsx"):
        return _read_xlsx_rows(content)
    if lowered.endswith(".csv") or lowered.endswith(".txt"):
        return _read_text_rows(content)
    if lowered.endswith(".xls"):
        raise ValueError("Eleventa a veces exporta .xls. Abrelo en Excel y guardalo como .xlsx o CSV.")
    raise ValueError("Formato no soportado. Usa archivo .xlsx o .csv exportado desde Eleventa.")


def map_eleventa_columns(headers: list[str]) -> dict[str, str]:
    normalized_headers = {_normalize_header(header): header for header in headers if header}
    mapping: dict[str, str] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if normalized_alias in normalized_headers:
                mapping[field] = normalized_headers[normalized_alias]
                break
    if "sku" not in mapping or "name" not in mapping:
        raise ValueError(
            "No se encontraron columnas obligatorias. Revisa que el archivo tenga Codigo y Descripcion."
        )
    return mapping


def _row_value(row: dict[str, str], column_name: str | None) -> str:
    if not column_name:
        return ""
    return (row.get(column_name) or "").strip()


def _get_or_create_department(db: Session, cache: dict[str, Department], name: str) -> tuple[Department | None, bool]:
    cleaned = name.strip()
    if not cleaned:
        return None, False
    key = cleaned.lower()
    if key in cache:
        return cache[key], False
    existing = db.query(Department).filter(Department.name.ilike(cleaned)).first()
    if existing:
        cache[key] = existing
        return existing, False
    department = Department(name=cleaned[:120], description="Importado desde Eleventa")
    db.add(department)
    db.flush()
    cache[key] = department
    return department, True


def _get_or_create_supplier(db: Session, cache: dict[str, Supplier], name: str) -> tuple[Supplier, bool]:
    cleaned = (name or DEFAULT_SUPPLIER_NAME).strip() or DEFAULT_SUPPLIER_NAME
    key = cleaned.lower()
    if key in cache:
        return cache[key], False
    existing = db.query(Supplier).filter(Supplier.name.ilike(cleaned)).first()
    if existing:
        cache[key] = existing
        return existing, False
    supplier = Supplier(name=cleaned[:200], notes="Creado automaticamente al importar desde Eleventa")
    db.add(supplier)
    db.flush()
    cache[key] = supplier
    return supplier, True


def import_eleventa_rows(
    db: Session,
    *,
    user_id: int,
    rows: list[dict[str, str]],
    column_map: dict[str, str],
    update_existing: bool = True,
    update_stock: bool = True,
    default_supplier_name: str = DEFAULT_SUPPLIER_NAME,
) -> EleventaImportStats:
    stats = EleventaImportStats()
    department_cache: dict[str, Department] = {}
    supplier_cache: dict[str, Supplier] = {}

    for index, row in enumerate(rows, start=2):
        sku = _row_value(row, column_map.get("sku"))
        name = _row_value(row, column_map.get("name"))
        if not sku and not name:
            continue
        if not sku:
            stats.skipped += 1
            stats.errors.append(f"Fila {index}: producto sin codigo (SKU).")
            continue
        if not name:
            stats.skipped += 1
            stats.errors.append(f"Fila {index}: producto {sku} sin descripcion.")
            continue

        sku = sku[:50]
        department_name = _row_value(row, column_map.get("department"))
        supplier_name = _row_value(row, column_map.get("supplier")) or default_supplier_name

        department, department_created = _get_or_create_department(db, department_cache, department_name)
        if department_created:
            stats.departments_created += 1

        supplier, supplier_created = _get_or_create_supplier(db, supplier_cache, supplier_name)
        if supplier_created:
            stats.suppliers_created += 1

        cost = _parse_number(_row_value(row, column_map.get("cost")))
        price = _parse_number(_row_value(row, column_map.get("price")))
        stock = _parse_number(_row_value(row, column_map.get("stock")))
        min_stock = _parse_number(_row_value(row, column_map.get("min_stock")))
        barcode = _normalize_barcode(_row_value(row, column_map.get("barcode")))

        product = db.query(Product).filter(Product.sku == sku).first()
        if product and not update_existing:
            stats.skipped += 1
            continue

        if product:
            product.name = name[:200]
            product.cost = max(cost, 0)
            product.price = max(price, 0)
            product.min_stock = max(min_stock, 0)
            product.supplier_id = supplier.id
            product.department_id = department.id if department else None
            if barcode:
                conflict = (
                    db.query(Product)
                    .filter(Product.barcode == barcode, Product.id != product.id)
                    .first()
                )
                if not conflict:
                    product.barcode = barcode
            if update_stock:
                before_stock = float(product.stock)
                product.stock = max(stock, 0)
                if product.stock != before_stock:
                    db.add(
                        InventoryMovement(
                            product_id=product.id,
                            created_by_user_id=user_id,
                            movement_type="adjustment",
                            quantity=round(product.stock - before_stock, 2),
                            before_stock=before_stock,
                            after_stock=product.stock,
                            notes="Importacion Eleventa",
                        )
                    )
            stats.updated += 1
            continue

        if barcode:
            conflict = db.query(Product).filter(Product.barcode == barcode).first()
            if conflict:
                barcode = None

        product = Product(
            sku=sku,
            barcode=barcode,
            name=name[:200],
            cost=max(cost, 0),
            price=max(price, 0),
            stock=max(stock, 0) if update_stock else 0,
            min_stock=max(min_stock, 0),
            supplier_id=supplier.id,
            department_id=department.id if department else None,
            tax_rate=0.12,
        )
        db.add(product)
        db.flush()
        if update_stock and product.stock > 0:
            db.add(
                InventoryMovement(
                    product_id=product.id,
                    created_by_user_id=user_id,
                    movement_type="entry",
                    quantity=product.stock,
                    before_stock=0,
                    after_stock=product.stock,
                    notes="Importacion Eleventa",
                )
            )
        stats.created += 1

    db.commit()
    return stats
